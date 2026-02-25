#!/usr/bin/env python3
"""
Script to analyze task 2 from the Excel file
"""

import sys
import os

# Try different libraries to read Excel
try:
    import pandas as pd

    def analyze_task():
        """Analyze task 2 using pandas"""
        file_path = os.path.join(os.path.dirname(__file__), "未来工作课题.xlsx")

        print(f"分析文件: {file_path}")
        print("=" * 60)

        # Read Excel file
        df = pd.read_excel(file_path)

        print(f"文件信息:")
        print(f"- 行数: {len(df)}")
        print(f"- 列数: {len(df.columns)}")
        print(f"- 列名: {df.columns.tolist()}")
        print()

        # Find task 2
        print("=== 序号2任务分析 ===")
        task_found = False
        target_task = 2

        for idx, row in df.iterrows():
            if isinstance(row.iloc[0], (int, float)) and row.iloc[0] == target_task:
                task_found = True
                print(f"序号: {row.iloc[0]}")
                if len(row) > 1 and pd.notna(row.iloc[1]):
                    print(f"标题: {row.iloc[1]}")
                if len(row) > 2 and pd.notna(row.iloc[2]):
                    print(f"背景: {row.iloc[2]}")
                if len(row) > 3 and pd.notna(row.iloc[3]):
                    print(f"任务: {row.iloc[3]}")
                # Check for additional columns
                if len(row) > 4:
                    for i, val in enumerate(row.iloc[4:], 5):
                        if pd.notna(val):
                            print(f"其他信息{i-4}: {val}")
                break

        if not task_found:
            print(f"未找到序号{target_task}的任务")
            # Print all rows to see what's available
            print("\n所有行数据:")
            for idx, row in df.iterrows():
                print(f"{idx+1}: {row.values.tolist()}")

except ImportError:
    try:
        import openpyxl

        def analyze_task():
            """Analyze task 2 using openpyxl"""
            file_path = os.path.join(os.path.dirname(__file__), "未来工作课题.xlsx")

            print(f"分析文件: {file_path}")
            print("=" * 60)

            # Load workbook
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            print(f"文件信息:")
            print(f"- 工作表: {sheet.title}")
            print(f"- 最大行数: {sheet.max_row}")
            print(f"- 最大列数: {sheet.max_column}")
            print()

            # Find task 2
            print("=== 序号2任务分析 ===")
            task_found = False
            target_task = 2

            for row in sheet.iter_rows(
                min_row=1, max_row=sheet.max_row, values_only=True
            ):
                if row and isinstance(row[0], (int, float)) and row[0] == target_task:
                    task_found = True
                    print(f"序号: {row[0]}")
                    if len(row) > 1 and row[1]:
                        print(f"标题: {row[1]}")
                    if len(row) > 2 and row[2]:
                        print(f"背景: {row[2]}")
                    if len(row) > 3 and row[3]:
                        print(f"任务: {row[3]}")
                    # Check for additional columns
                    if len(row) > 4:
                        for i, val in enumerate(row[4:], 5):
                            if val:
                                print(f"其他信息{i-4}: {val}")
                    break

            if not task_found:
                print(f"未找到序号{target_task}的任务")
                # Print all rows to see what's available
                print("\n所有行数据:")
                for i, row in enumerate(
                    sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True),
                    1,
                ):
                    print(f"{i}: {row}")

    except ImportError:
        print("请安装必要的库: pip install pandas openpyxl")
        sys.exit(1)

if __name__ == "__main__":
    analyze_task()
#!/usr/bin/env python3
"""
Script to analyze task 2 from the Excel file
"""

import sys
import os

# Try different libraries to read Excel
try:
    import pandas as pd

    def analyze_task():
        """Analyze task 2 using pandas"""
        file_path = os.path.join(os.path.dirname(__file__), "未来工作课题.xlsx")

        print(f"分析文件: {file_path}")
        print("=" * 60)

        # Read Excel file
        df = pd.read_excel(file_path)

        print(f"文件信息:")
        print(f"- 行数: {len(df)}")
        print(f"- 列数: {len(df.columns)}")
        print(f"- 列名: {df.columns.tolist()}")
        print()

        # Find task 2
        print("=== 序号2任务分析 ===")
        task_found = False
        target_task = 2

        for idx, row in df.iterrows():
            if isinstance(row.iloc[0], (int, float)) and row.iloc[0] == target_task:
                task_found = True
                print(f"序号: {row.iloc[0]}")
                if len(row) > 1 and pd.notna(row.iloc[1]):
                    print(f"标题: {row.iloc[1]}")
                if len(row) > 2 and pd.notna(row.iloc[2]):
                    print(f"背景: {row.iloc[2]}")
                if len(row) > 3 and pd.notna(row.iloc[3]):
                    print(f"任务: {row.iloc[3]}")
                # Check for additional columns
                if len(row) > 4:
                    for i, val in enumerate(row.iloc[4:], 5):
                        if pd.notna(val):
                            print(f"其他信息{i-4}: {val}")
                break

        if not task_found:
            print(f"未找到序号{target_task}的任务")
            # Print all rows to see what's available
            print("\n所有行数据:")
            for idx, row in df.iterrows():
                print(f"{idx+1}: {row.values.tolist()}")

except ImportError:
    try:
        import openpyxl

        def analyze_task():
            """Analyze task 2 using openpyxl"""
            file_path = os.path.join(os.path.dirname(__file__), "未来工作课题.xlsx")

            print(f"分析文件: {file_path}")
            print("=" * 60)

            # Load workbook
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            print(f"文件信息:")
            print(f"- 工作表: {sheet.title}")
            print(f"- 最大行数: {sheet.max_row}")
            print(f"- 最大列数: {sheet.max_column}")
            print()

            # Find task 2
            print("=== 序号2任务分析 ===")
            task_found = False
            target_task = 2

            for row in sheet.iter_rows(
                min_row=1, max_row=sheet.max_row, values_only=True
            ):
                if row and isinstance(row[0], (int, float)) and row[0] == target_task:
                    task_found = True
                    print(f"序号: {row[0]}")
                    if len(row) > 1 and row[1]:
                        print(f"标题: {row[1]}")
                    if len(row) > 2 and row[2]:
                        print(f"背景: {row[2]}")
                    if len(row) > 3 and row[3]:
                        print(f"任务: {row[3]}")
                    # Check for additional columns
                    if len(row) > 4:
                        for i, val in enumerate(row[4:], 5):
                            if val:
                                print(f"其他信息{i-4}: {val}")
                    break

            if not task_found:
                print(f"未找到序号{target_task}的任务")
                # Print all rows to see what's available
                print("\n所有行数据:")
                for i, row in enumerate(
                    sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True),
                    1,
                ):
                    print(f"{i}: {row}")

    except ImportError:
        print("请安装必要的库: pip install pandas openpyxl")
        sys.exit(1)

if __name__ == "__main__":
    analyze_task()
