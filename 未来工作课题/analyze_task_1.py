#!/usr/bin/env python3
"""
Script to analyze task 1 and recent papers from the Excel file
"""

import sys
import os

# Try different libraries to read Excel
try:
    import pandas as pd

    def analyze_excel():
        """Analyze Excel file using pandas"""
        file_path = os.path.join(os.path.dirname(__file__), "未来工作课题.xlsx")

        print(f"分析文件: {file_path}")
        print("=" * 60)

        # Read Excel file
        df = pd.read_excel(file_path)

        print(f"文件信息:")
        print(f"- 行数: {len(df)}")
        print(f"- 列数: {len(df.columns)}")
        print(f"- 列名: {df.columns.tolist()}")
        print()

        # Find task 1
        print("=== 序号1标题背景任务分析 ===")
        task_1_found = False
        for idx, row in df.iterrows():
            if isinstance(row.iloc[0], (int, float)) and row.iloc[0] == 1:
                task_1_found = True
                print(f"序号: {row.iloc[0]}")
                if len(row) > 1 and pd.notna(row.iloc[1]):
                    print(f"标题: {row.iloc[1]}")
                if len(row) > 2 and pd.notna(row.iloc[2]):
                    print(f"背景任务: {row.iloc[2]}")
                if len(row) > 3:
                    for i, val in enumerate(row.iloc[3:], 4):
                        if pd.notna(val):
                            print(f"其他信息{i-3}: {val}")
                break

        if not task_1_found:
            print("未找到序号1的任务")

        print()
        print("=== 最近3年涉及的论文 ===")

        # Search for papers section
        papers_found = False
        for idx, row in df.iterrows():
            for col in df.columns:
                cell_val = row[col]
                if isinstance(cell_val, str) and "论文" in cell_val:
                    papers_found = True
                    print(f"论文部分: {cell_val}")
                    print()

                    # Print next few rows as papers
                    for i in range(idx + 1, min(idx + 20, len(df))):
                        paper_row = df.iloc[i]
                        if any(pd.notna(val) for val in paper_row):
                            print(f"{i-idx}: {paper_row.values.tolist()}")
                        else:
                            break
                    break
            if papers_found:
                break

        if not papers_found:
            print("未找到论文部分")

except ImportError:
    try:
        import openpyxl

        def analyze_excel():
            """Analyze Excel file using openpyxl"""
            file_path = os.path.join(os.path.dirname(__file__), "未来工作课题.xlsx")

            print(f"分析文件: {file_path}")
            print("=" * 60)

            # Load workbook
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            print(f"文件信息:")
            print(f"- 工作表: {sheet.title}")
            print(f"- 最大行数: {sheet.max_row}")
            print(f"- 最大列数: {sheet.max_column}")
            print()

            # Find task 1
            print("=== 序号1标题背景任务分析 ===")
            task_1_found = False
            for row in sheet.iter_rows(
                min_row=1, max_row=sheet.max_row, values_only=True
            ):
                if row and isinstance(row[0], (int, float)) and row[0] == 1:
                    task_1_found = True
                    print(f"序号: {row[0]}")
                    if len(row) > 1 and row[1]:
                        print(f"标题: {row[1]}")
                    if len(row) > 2 and row[2]:
                        print(f"背景任务: {row[2]}")
                    if len(row) > 3:
                        for i, val in enumerate(row[3:], 4):
                            if val:
                                print(f"其他信息{i-3}: {val}")
                    break

            if not task_1_found:
                print("未找到序号1的任务")

            print()
            print("=== 最近3年涉及的论文 ===")

            # Search for papers section
            papers_found = False
            for i, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), 1
            ):
                if row:
                    for val in row:
                        if isinstance(val, str) and "论文" in val:
                            papers_found = True
                            print(f"论文部分: {val}")
                            print()

                            # Print next few rows as papers
                            for j in range(i + 1, min(i + 20, sheet.max_row + 1)):
                                paper_row = next(
                                    sheet.iter_rows(
                                        min_row=j, max_row=j, values_only=True
                                    )
                                )[0]
                                if paper_row:
                                    print(f"{j-i}: {paper_row}")
                                else:
                                    break
                            break
                    if papers_found:
                        break

            if not papers_found:
                print("未找到论文部分")

    except ImportError:
        print("请安装必要的库: pip install pandas openpyxl")
        sys.exit(1)

if __name__ == "__main__":
    analyze_excel()
